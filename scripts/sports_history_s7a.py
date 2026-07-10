#!/usr/bin/env python3
"""sports_history_s7a.py — Q4/S7a: historical sourcing for the sports-CLV backtest.

S7's binding test (S7b/S7c, later stages) needs a *season* of Kalshi moneyline
history at decision-time real asks, matched to a de-vigged sharp closing line. This
stage only SOURCES that data and documents its provenance — no backtest math here.

================================================================================
WHAT WORKED: World Cup 2026 (Kalshi KXWCGAME + football-data.co.uk)
================================================================================
Kalshi's public `/events` endpoint (status=settled, with_nested_markets=true) returns
every FINALIZED KXWCGAME event with its outcome markets' own `result` (yes/no) and
`settlement_value_dollars` inline — no extra call needed for the settlement leg. As of
this run: 97 completed World Cup games (2026-06-11 .. current), 291 outcome markets.
For each outcome market we additionally pull the FULL hourly candlestick series
(`/series/{s}/markets/{t}/candlesticks`, `period_interval=60`) spanning the market's
open_time..close_time — Kalshi's own published `yes_ask` OHLC, i.e. `real_ask` by
construction (same complement-of-best-NO-bid taker price as every other real_ask site
in this repo). We persist the FULL candlestick series rather than picking one
"decision time" now, so S7b can choose its decision-time definition without a second
network pass.

The free historical closing-odds source is football-data.co.uk's public
`WorldCup2026.xlsx` (no key, no auth, updated as the tournament progresses) — average
closing decimal odds across their tracked bookmaker panel (`H-Avg`/`D-Avg`/`A-Avg`
columns) for every match, home/draw/away. This is NOT Pinnacle specifically (Q1's
preferred single sharp book) — it is a multi-book average, a reasonable but weaker
sharp-consensus proxy. De-vigged via core.odds (decimal odds -> implied prob ->
multiplicative de-vig), tagged `synthetic` per CLAUDE.md (a de-vig is a model, not a
fill). Team names differ cosmetically between the two sources (Kalshi "IR Iran" vs
football-data "Iran", etc.) — TEAM_NAME_ALIASES documents every observed discrepancy;
an unmapped name is a genuine unmatched game, not a silent alias.

================================================================================
WHAT DID NOT (FULLY) WORK: last-season NFL/NBA
================================================================================
Verified directly against the live API this run (all reproducible by re-running this
script's `probe_last_season_availability`, whose result is logged in every run's
summary): Kalshi's public `/markets` listing appears to retain settled/finalized
markets for a RECENT WINDOW ONLY, not the full history of a series.
  - **KXNFLGAME**: `status=settled`/`status=closed` return ZERO rows. Unfiltered
    `/markets?series_ticker=KXNFLGAME` returns only 66 markets, ALL with close_time in
    2026-08/09 (next season's not-yet-played preseason/openers) — the 2025 season that
    finished in February 2026 (~5 months ago) has fully aged out. Without a market
    ticker we cannot even call candlesticks (404s on an unknown ticker) — a genuine
    data-availability wall, not a code bug.
  - **KXNBAGAME**: `status=settled` returns 72 outcome markets across 36 games, but
    ONLY the tail of the playoffs (2026-05-05 .. 2026-06-14 — conference finals through
    the Finals). The regular season (Oct 2025-Apr 2026) is gone the same way NFL's
    full season is. So NBA is a partial win, not a full one: a real, sourceable
    36-game playoff dataset exists (candlesticks fetchable, real_ask), but no matched
    free historical odds leg is sourced for it in this run — that plus the missing
    regular season are flagged for a future stage, not solved here.
A full season of NFL/NBA backtest data needs a DIFFERENT source (Kalshi's
historical-data archive product, if any, or a third-party historical-odds+result
feed). S7b/S7c should run on the World Cup dataset first (this run's deliverable);
the NBA playoff window is a documented opportunity for a follow-up stage.

Run:
    python -m scripts.sports_history_s7a                  # live pass, writes tape/
    python -m scripts.sports_history_s7a --limit 5         # cap games (smoke test)
"""
from __future__ import annotations

import argparse
import io
import re
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import requests

from core.canonical import canonical_json, sha256_hex
from core.io import REPO_ROOT
from core.odds import decimal_to_implied_prob, devig_multiplicative
from core.timeutil import _parse_iso
from validation.v3_market import Kalshi, _load_venue_cfg

STORE = REPO_ROOT / "tape" / "sports_history_s7"
FOOTBALL_DATA_WC_URL = "https://www.football-data.co.uk/WorldCup2026.xlsx"
FOOTBALL_DATA_SHEET = "WorldCup2026"
SCHEMA_VERSION = "sports_history_s7.v0"
CANDLE_PERIOD_INTERVAL_MIN = 60
# World Cup outcome markets are listed (open_time) as early as tournament announcement —
# empirically up to ~140 days before their game — but any decision-time definition S7b
# is likely to pick (T-24h, T-1h, ...) only needs the pre-game window. Capped at 7 days
# before close_time so the tape stays proportionate; the cap is explicit and logged
# (`candle_window_truncated`) per record, never a silent drop.
CANDLE_LOOKBACK_HOURS = 24 * 7

# Kalshi display name -> football-data.co.uk display name, for every observed mismatch.
# An unmapped-but-actually-different name shows up as a genuine unmatched game (logged),
# never a silent guess.
TEAM_NAME_ALIASES = {
    "Bosnia and Herzegovina": "Bosnia & Herzegovina",
    "Congo DR": "D.R. Congo",
    "Czechia": "Czech Republic",
    "IR Iran": "Iran",
    "Korea Republic": "South Korea",
    "Turkiye": "Turkey",
}

_EVENT_WINNER_TITLE_RE = re.compile(r"^(.+) vs (.+) Winner\?$")


def _slug(name: str) -> str:
    canon = TEAM_NAME_ALIASES.get(name, name)
    return re.sub(r"[^a-z0-9]", "", canon.lower())


# --------------------------------------------------------------------------- #
# Kalshi leg — settled events, real_ask candlesticks
# --------------------------------------------------------------------------- #
def parse_event_teams(nested_markets: List[dict]) -> Optional[Tuple[str, str]]:
    """The nested market titled '<Home> vs <Away> Winner?' names the two teams. Kalshi's
    event-level `title` ("<A> vs <B>: Regulation Time Moneyline") is not reliably in the
    same word order/format, so we parse the per-outcome market title instead — and we
    must search for it (it is not always the first nested market)."""
    for m in nested_markets:
        match = _EVENT_WINNER_TITLE_RE.match(m.get("title", "") or "")
        if match:
            return match.group(1).strip(), match.group(2).strip()
    return None


def _outcome_side(market: dict) -> str:
    sub = (market.get("yes_sub_title") or "").strip().lower()
    if sub.endswith("tie"):
        return "tie"
    return "home_or_away"  # disambiguated by the caller via team-name match


def probe_last_season_availability(client, series_tickers: List[str]) -> Dict[str, Dict]:
    """Re-runnable check: does `/markets?series_ticker=...&status=settled` return ANY
    rows for these series? Documents the NFL/NBA finding above without hardcoding an
    assumption — re-run this after the retention window has moved to see if it changed."""
    out: Dict[str, Dict] = {}
    for st in series_tickers:
        try:
            settled = client.markets(st, status="settled", limit=200)
            closed = client.markets(st, status="closed", limit=200)
        except Exception as exc:
            out[st] = {"error": str(exc)}
            continue
        out[st] = {"n_settled": len(settled), "n_closed": len(closed)}
    return out


def fetch_worldcup_games(client, series: str = "KXWCGAME", limit: Optional[int] = None
                         ) -> Tuple[List[dict], List[str]]:
    """Every finalized KXWCGAME event, nested markets included. Returns (games, warnings)
    where a game with no parseable '<A> vs <B> Winner?' outcome is skipped and logged as
    a warning rather than silently dropped."""
    events = client.events(series, status="settled", limit=200)
    games: List[dict] = []
    warnings: List[str] = []
    for ev in events:
        markets = ev.get("markets") or []
        teams = parse_event_teams(markets)
        if teams is None:
            warnings.append(f"{ev.get('event_ticker')}: no parseable '<A> vs <B> Winner?' "
                            f"outcome market, skipped")
            continue
        games.append({"event": ev, "home": teams[0], "away": teams[1]})
        if limit and len(games) >= limit:
            break
    return games, warnings


def candle_window(open_time: str, close_time: str) -> Tuple[int, int, bool]:
    """(start_ts, end_ts, truncated) for a market's candlestick fetch: the full
    open_time..close_time span, capped to the last CANDLE_LOOKBACK_HOURS before close
    (see the constant's comment) — `truncated` says whether the cap actually bit."""
    open_ts = int(_parse_iso(open_time).timestamp())
    close_ts = int(_parse_iso(close_time).timestamp())
    lookback_start = close_ts - CANDLE_LOOKBACK_HOURS * 3600
    start_ts = max(open_ts, lookback_start)
    return start_ts, close_ts + 3600, start_ts > open_ts


def fetch_outcome_candles(client, series: str, market: dict) -> Tuple[Optional[List[dict]], bool]:
    """Hourly candlestick series over `candle_window`. Returns (candles, truncated); candles
    is None on fetch failure (an outcome we could not get candles for is dropped from
    that outcome's record, never faked as an empty-but-successful series)."""
    ticker = market.get("ticker", "")
    try:
        start_ts, end_ts, truncated = candle_window(market["open_time"], market["close_time"])
        candles = client.candlesticks(series, ticker, CANDLE_PERIOD_INTERVAL_MIN, start_ts, end_ts)
        return candles, truncated
    except Exception:
        return None, False


# --------------------------------------------------------------------------- #
# odds leg — football-data.co.uk free public closing-odds average (synthetic)
# --------------------------------------------------------------------------- #
def fetch_football_data_bytes(url: str = FOOTBALL_DATA_WC_URL, timeout: float = 30.0) -> bytes:
    r = requests.get(url, timeout=timeout)
    r.raise_for_status()
    return r.content


def load_worldcup_odds_rows(xlsx_bytes: bytes, sheet: str = FOOTBALL_DATA_SHEET) -> List[Dict[str, Any]]:
    """Parse football-data.co.uk's WorldCup2026.xlsx into per-match rows. Pure function
    of the file bytes — no network, so this is independently unit-testable against a
    small fixture workbook."""
    import openpyxl  # optional-dep (analysis extra); imported lazily so the base
                      # substrate + invariants keep running without it installed.

    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    ws = wb[sheet]
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {name: i for i, name in enumerate(header)}
    required = ("Home", "Away", "Date", "H-Avg", "D-Avg", "A-Avg")
    missing = [c for c in required if c not in idx]
    if missing:
        raise ValueError(f"WorldCup2026.xlsx missing expected columns: {missing}")

    rows: List[Dict[str, Any]] = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        home, away = r[idx["Home"]], r[idx["Away"]]
        date_val = r[idx["Date"]]
        if not home or not away or date_val is None:
            continue
        rows.append({
            "home": str(home), "away": str(away),
            "date": date_val.date().isoformat() if hasattr(date_val, "date") else str(date_val),
            "h_avg": r[idx["H-Avg"]], "d_avg": r[idx["D-Avg"]], "a_avg": r[idx["A-Avg"]],
        })
    return rows


def match_odds_row(rows: List[Dict[str, Any]], home: str, away: str) -> Optional[Dict[str, Any]]:
    """Order-agnostic team-pair match (football-data and Kalshi don't always agree on
    which side is 'home'). A World Cup pairing meets at most once except in a repeat
    final/3rd-place edge case, which we do not expect in this dataset; ties are broken
    by taking the first match (logged as ambiguous is out of scope for this stage)."""
    key = frozenset({_slug(home), _slug(away)})
    for row in rows:
        if frozenset({_slug(row["home"]), _slug(row["away"])}) == key:
            return row
    return None


def devig_odds_row(row: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    """H-Avg/D-Avg/A-Avg decimal odds -> de-vigged fair probs (sums to 1.0), tagged
    synthetic. None if any leg is missing (e.g. a match still in progress)."""
    h, d, a = row.get("h_avg"), row.get("d_avg"), row.get("a_avg")
    if h is None or d is None or a is None:
        return None
    implied = [decimal_to_implied_prob(float(x)) for x in (h, d, a)]
    fair_home, fair_draw, fair_away = devig_multiplicative(implied)
    return {
        "home_odds_decimal": float(h), "draw_odds_decimal": float(d), "away_odds_decimal": float(a),
        "fair_home": fair_home, "fair_draw": fair_draw, "fair_away": fair_away,
        "price_source_tag": "synthetic",
        "odds_source": "football-data.co.uk WorldCup2026.xlsx (H-Avg/D-Avg/A-Avg, "
                       "multi-book average — not Pinnacle-specific)",
    }


# --------------------------------------------------------------------------- #
# assembly — one record per game
# --------------------------------------------------------------------------- #
def build_game_record(game: Dict[str, Any], client, series: str,
                      odds_rows: List[Dict[str, Any]], fetch_ts: str) -> Dict[str, Any]:
    ev = game["event"]
    home, away = game["home"], game["away"]
    markets = ev.get("markets") or []

    outcomes: List[Dict[str, Any]] = []
    n_candle_failures = 0
    for m in markets:
        candles, truncated = fetch_outcome_candles(client, series, m)
        if candles is None:
            n_candle_failures += 1
        outcomes.append({
            "market_ticker": m.get("ticker", ""),
            "yes_sub_title": m.get("yes_sub_title", ""),
            "result": m.get("result", ""),
            "settlement_value_dollars": m.get("settlement_value_dollars", ""),
            "open_time": m.get("open_time", ""),
            "close_time": m.get("close_time", ""),
            "candles": candles if candles is not None else [],
            "candle_fetch_ok": candles is not None,
            "candle_window_truncated": truncated,
        })

    odds_row = match_odds_row(odds_rows, home, away)
    odds_devig = devig_odds_row(odds_row) if odds_row else None

    raw_payload = {"event": ev, "outcomes": outcomes}
    return {
        "schema_version": SCHEMA_VERSION,
        "sport": "soccer_world_cup_2026",
        "kalshi_event_ticker": ev.get("event_ticker", ""),
        "kalshi_series": series,
        "home_team": home, "away_team": away,
        "outcomes": outcomes,
        "n_outcomes": len(outcomes),
        "n_candle_fetch_failures": n_candle_failures,
        "price_source_tag_kalshi": "real_ask",
        "kalshi_raw_sha256": sha256_hex(canonical_json(raw_payload)),
        "odds_match": {"matched": odds_row is not None, **(odds_devig or {})},
        "fetch_ts": fetch_ts,
    }


def run(limit: Optional[int] = None, min_interval: float = 0.2,
        client=None, store: Optional[Path] = None,
        odds_bytes_fetcher=None) -> Dict[str, Any]:
    """One historical-sourcing pass. `client`/`store`/`odds_bytes_fetcher` are injectable
    for offline testing; production defaults to the live Kalshi client, the real tape
    store, and a live fetch of football-data.co.uk's public xlsx."""
    store = Path(store) if store is not None else STORE
    if client is None:
        cfg = _load_venue_cfg()
        client = Kalshi(cfg["api_base"], min_interval=min_interval)
    if odds_bytes_fetcher is None:
        odds_bytes_fetcher = fetch_football_data_bytes

    cap_ts = datetime.now(timezone.utc)
    fetch_ts = cap_ts.isoformat()
    run_id = cap_ts.strftime("%Y%m%dT%H%M%SZ")

    availability = probe_last_season_availability(client, ["KXNFLGAME", "KXNBAGAME"])

    games, parse_warnings = fetch_worldcup_games(client, limit=limit)

    odds_bytes = odds_bytes_fetcher()
    odds_raw_sha256 = sha256_hex(odds_bytes)
    odds_rows = load_worldcup_odds_rows(odds_bytes)

    records = [build_game_record(g, client, "KXWCGAME", odds_rows, fetch_ts) for g in games]

    store.mkdir(parents=True, exist_ok=True)
    out_path = store / "worldcup2026.jsonl"
    with open(out_path, "a") as f:
        for rec in records:
            rec["odds_raw_sha256"] = odds_raw_sha256
            rec["run_id"] = run_id
            f.write(canonical_json(rec) + "\n")
    (store / f"worldcup2026-odds-source-{run_id}.xlsx").write_bytes(odds_bytes)

    n_matched = sum(1 for r in records if r["odds_match"]["matched"])
    n_candle_fail = sum(r["n_candle_fetch_failures"] for r in records)
    summary = {
        "run_id": run_id, "fetch_ts": fetch_ts,
        "n_games": len(records),
        "n_odds_matched": n_matched,
        "n_odds_unmatched": len(records) - n_matched,
        "n_candle_fetch_failures": n_candle_fail,
        "n_parse_warnings": len(parse_warnings),
        "last_season_nfl_nba_availability": availability,
        "out_path": str(out_path),
    }
    print(f"[sports_history_s7a] {summary['n_games']} World Cup games, "
          f"{n_matched}/{summary['n_games']} odds-matched, "
          f"{n_candle_fail} candlestick fetch failures, "
          f"NFL/NBA settled availability: {availability} -> {out_path}")
    if parse_warnings:
        print(f"[sports_history_s7a] WARN {len(parse_warnings)} unparseable events "
              f"skipped: {parse_warnings}", file=sys.stderr)
    return summary


def main(argv: Optional[List[str]] = None) -> int:
    ap = argparse.ArgumentParser(description="S7a historical sourcing (read-only)")
    ap.add_argument("--limit", type=int, default=None, help="cap games fetched (testing)")
    ap.add_argument("--min-interval", type=float, default=0.2)
    args = ap.parse_args(argv)
    run(limit=args.limit, min_interval=args.min_interval)
    return 0


if __name__ == "__main__":
    sys.exit(main())
